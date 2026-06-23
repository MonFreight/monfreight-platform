"""
Mon Freight Customer Data Platform — internal tool for recording shipments,
printing customs labels, and exporting Air Cargo / Labels Excel workbooks.

Run locally:
    pip install -r requirements.txt
    uvicorn app:app --reload
"""
from __future__ import annotations

import ast
import datetime as dt
import io
import operator
import os
from pathlib import Path
from typing import Any, Optional

from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from sqlalchemy import (Boolean, Column, Date, DateTime, Float, Integer,
                        String, create_engine, select, text)
from sqlalchemy.orm import DeclarativeBase, Session

from label_excel import (build_aircargo_xlsx, build_labels_xlsx,
                         _format_receiver_address, _strip_quantities,
                         PAGE_HEIGHT as LABEL_PAGE_HEIGHT)
from xlsx_html import render_print_page

# --------------------------------------------------------------------------
# config
# --------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"
DATA_DIR.mkdir(exist_ok=True)

# Seed the label template into the data dir on first boot. Needed on hosts
# (e.g. Railway) where an empty persistent volume is mounted over data/.
_tpl_seed = BASE_DIR / "assets" / "label_template.xlsx"
_tpl_dest = DATA_DIR / "label_template.xlsx"
if _tpl_seed.exists() and not _tpl_dest.exists():
    import shutil as _shutil
    _shutil.copy(_tpl_seed, _tpl_dest)

DB_PATH = os.environ.get("DB_PATH", str(DATA_DIR / "monfreight.db"))
DB_URL = os.environ.get("DATABASE_URL", f"sqlite:///{DB_PATH}")
if DB_URL.startswith("postgres://"):
    DB_URL = DB_URL.replace("postgres://", "postgresql+psycopg://", 1)

engine = create_engine(
    DB_URL,
    connect_args={"check_same_thread": False} if DB_URL.startswith("sqlite") else {},
)


# --------------------------------------------------------------------------
# models
# --------------------------------------------------------------------------
class Base(DeclarativeBase):
    pass


class Shipment(Base):
    __tablename__ = "shipments"
    id = Column(Integer, primary_key=True)
    batch_date = Column(Date, nullable=False, index=True)
    box_number = Column(Integer, nullable=False)
    mf_number = Column(String, nullable=False, unique=True)

    sender_name = Column(String, default="")
    sender_phone = Column(String, default="")
    sender_address = Column(String, default="")
    sender_city = Column(String, default="")
    sender_country = Column(String, default="Австрали")
    sender_postal = Column(String, default="")

    receiver_name = Column(String, default="")
    receiver_phone = Column(String, default="")
    receiver_address = Column(String, default="")
    receiver_city = Column(String, default="")
    receiver_country = Column(String, default="Монгол")

    description = Column(String, default="")
    declared_value = Column(Float, default=0.0)
    weight = Column(Float, default=0.0)
    price_formula = Column(String, default="")
    price_aud = Column(Float, default=0.0)
    extra_charges = Column(Float, default=0.0)
    total_aud = Column(Float, default=0.0)
    paid = Column(Boolean, default=False)
    delivery_note = Column(String, default="")
    notes = Column(String, default="")
    link_group = Column(Integer, nullable=True)
    package_id = Column(Integer, nullable=True, index=True)

    created_at = Column(DateTime, default=dt.datetime.now)


Base.metadata.create_all(engine)


def _ensure_columns():
    """Lightweight migration: add new columns to old SQLite databases.
    Resilient — one failed ALTER doesn't stop the rest."""
    if not DB_URL.startswith("sqlite"):
        return
    cols_to_add = [
        ("paid", "ALTER TABLE shipments ADD COLUMN paid BOOLEAN DEFAULT 0"),
        ("price_formula", "ALTER TABLE shipments ADD COLUMN price_formula TEXT DEFAULT ''"),
        ("total_aud", "ALTER TABLE shipments ADD COLUMN total_aud FLOAT DEFAULT 0"),
        ("total_override", "ALTER TABLE shipments ADD COLUMN total_override BOOLEAN DEFAULT 0"),
        ("extra_charges", "ALTER TABLE shipments ADD COLUMN extra_charges FLOAT DEFAULT 0"),
        ("link_group",    "ALTER TABLE shipments ADD COLUMN link_group INTEGER DEFAULT NULL"),
        ("package_id",    "ALTER TABLE shipments ADD COLUMN package_id INTEGER DEFAULT NULL"),
    ]
    with engine.connect() as conn:
        try:
            existing = {r[1] for r in conn.execute(
                text("PRAGMA table_info(shipments)")).fetchall()}
        except Exception:
            existing = set()
        for col, ddl in cols_to_add:
            if col in existing:
                continue
            try:
                conn.execute(text(ddl))
                conn.commit()
                print(f"[migration] added column: {col}")
            except Exception as e:
                print(f"[migration] could not add column {col}: {e}")


_ensure_columns()


def _resync_mf_numbers() -> None:
    """One-shot data fix: rebuild every shipment's mf_number from its
    (batch_date, box_number) so they always match — i.e. BOX 1 →
    MFYYMMDD001, BOX 2 → MFYYMMDD002, etc. Runs at startup, idempotent.
    Two-phase update (placeholder → real) so the unique constraint on
    mf_number doesn't reject a swap mid-flight."""
    fixed = []
    skipped = []
    with Session(engine) as s:
        ships = list(s.scalars(select(Shipment)).all())
        targets = []
        for sh in ships:
            expected = mf_for(sh.batch_date, sh.box_number)
            if sh.mf_number != expected:
                targets.append((sh, expected))
        if not targets:
            return
        for sh, _ in targets:
            sh.mf_number = f"__resync_{sh.id}__"
        s.flush()
        for sh, expected in targets:
            existing = s.scalar(
                select(Shipment).where(
                    Shipment.mf_number == expected,
                    Shipment.id != sh.id,
                )
            )
            if existing:
                skipped.append(
                    f"id={sh.id} batch={sh.batch_date} box={sh.box_number} "
                    f"would collide with {expected}")
                sh.mf_number = f"DUP_{sh.id}_{expected}"
            else:
                sh.mf_number = expected
                fixed.append(sh.id)
        s.commit()
    if fixed:
        print(f"[migration] rebuilt MF numbers for {len(fixed)} shipment(s)")
    for msg in skipped:
        print(f"[migration] WARN: {msg}")


# --------------------------------------------------------------------------
# safe formula evaluator
# --------------------------------------------------------------------------
class FormulaError(ValueError):
    pass


_OPS = {
    ast.Add: operator.add,
    ast.Sub: operator.sub,
    ast.Mult: operator.mul,
    ast.Div: operator.truediv,
    ast.USub: operator.neg,
    ast.UAdd: operator.pos,
    ast.Mod: operator.mod,
    # ast.Pow intentionally omitted — `weight ** 999999` could DoS the CPU.
}
_ALLOWED_NAMES = {"weight", "value", "w", "v"}


def evaluate_price(text_in: str, *, weight: float = 0.0,
                   declared_value: float = 0.0) -> float:
    """Evaluate a price expression. Plain numbers, or formulas referring to
    `weight` and `value` (or short aliases `w` and `v`).

    Examples:
        '25'                  -> 25
        '=weight * 5'         -> weight * 5
        '=weight*5+10'        -> weight*5 + 10
        '=w*7 + v*0.05'       -> weight*7 + value*0.05
    """
    if text_in is None:
        return 0.0
    s = str(text_in).strip()
    if not s:
        return 0.0
    if s.startswith("="):
        s = s[1:]

    try:
        tree = ast.parse(s, mode="eval")
    except SyntaxError as e:
        raise FormulaError(f"Could not parse formula: {e.msg}") from e

    env = {"weight": float(weight or 0), "value": float(declared_value or 0),
           "w": float(weight or 0), "v": float(declared_value or 0)}

    def visit(node: ast.AST) -> float:
        if isinstance(node, ast.Expression):
            return visit(node.body)
        if isinstance(node, ast.Constant):
            if isinstance(node.value, (int, float)):
                return float(node.value)
            raise FormulaError(f"Only numbers allowed; got {node.value!r}")
        if isinstance(node, ast.Num):  # py<3.8 fallback
            return float(node.n)
        if isinstance(node, ast.Name):
            if node.id not in _ALLOWED_NAMES:
                raise FormulaError(f"Unknown variable {node.id!r}; "
                                    "use 'weight' or 'value'")
            return env[node.id]
        if isinstance(node, ast.BinOp):
            op = _OPS.get(type(node.op))
            if not op:
                raise FormulaError(f"Operator {type(node.op).__name__} not allowed")
            return op(visit(node.left), visit(node.right))
        if isinstance(node, ast.UnaryOp):
            op = _OPS.get(type(node.op))
            if not op:
                raise FormulaError(f"Operator {type(node.op).__name__} not allowed")
            return op(visit(node.operand))
        raise FormulaError(f"Unsupported expression: {ast.dump(node)}")

    try:
        return float(visit(tree))
    except ZeroDivisionError:
        raise FormulaError("Division by zero")


# --------------------------------------------------------------------------
# pydantic schemas
# --------------------------------------------------------------------------
class ShipmentIn(BaseModel):
    batch_date: dt.date
    box_number: Optional[int] = None  # None → auto-assign next available
    sender_name: str = ""
    sender_phone: str = ""
    sender_address: str = ""
    sender_city: str = ""
    sender_country: str = "Австрали"
    sender_postal: str = ""
    receiver_name: str = ""
    receiver_phone: str = ""
    receiver_address: str = ""
    receiver_city: str = ""
    receiver_country: str = "Монгол"
    description: str = ""
    declared_value: float = 0.0
    weight: float = 0.0
    price_formula: str = ""
    price_aud: float = 0.0
    extra_charges: float = 0.0
    paid: bool = False
    delivery_note: str = ""
    notes: str = ""


class IdsPayload(BaseModel):
    ids: list[int]


class BulkCreateIn(BaseModel):
    """Create several blank shipments at once for a given batch date.
    Each gets an auto-assigned box number + MF number; all other fields use
    their model defaults so staff can fill in the details later."""
    batch_date: dt.date
    count: int = 1


class ShipmentPatch(BaseModel):
    """Partial update for inline edits."""
    batch_date: Optional[dt.date] = None
    box_number: Optional[int] = None
    paid: Optional[bool] = None
    price_formula: Optional[str] = None
    weight: Optional[float] = None
    declared_value: Optional[float] = None
    description: Optional[str] = None
    delivery_note: Optional[str] = None
    extra_charges: Optional[float] = None


# --------------------------------------------------------------------------
# helpers
# --------------------------------------------------------------------------
def next_box_for(session: Session, batch_date: dt.date) -> int:
    rows = session.scalars(
        select(Shipment.box_number).where(Shipment.batch_date == batch_date)
    ).all()
    return (max(rows) + 1) if rows else 1


def mf_for(batch_date: dt.date, box_number: int) -> str:
    """MFYYMMDDNNN — e.g. MF260426001."""
    return f"MF{batch_date.strftime('%y%m%d')}{box_number:03d}"


# Now that `mf_for` exists, run the one-shot data migration that ensures
# every existing shipment's MF number matches its (batch_date, box_number).
_resync_mf_numbers()


def _username(request: Request) -> str:
    """Extract the signed-in username from the request state (set by auth middleware)."""
    user = getattr(request.state, "user", None)
    return user.get("u", "unknown") if user else "unknown"


def _client_ip(request: Request) -> str:
    return request.client.host if request.client else ""


def shipments_for(session: Session, batch_date: dt.date) -> list[Shipment]:
    return list(session.scalars(
        select(Shipment)
        .where(Shipment.batch_date == batch_date)
        .order_by(Shipment.box_number)
    ).all())


def all_shipments(session: Session,
                   start: Optional[dt.date] = None,
                   end: Optional[dt.date] = None) -> list[Shipment]:
    q = select(Shipment).order_by(Shipment.batch_date.desc(),
                                   Shipment.box_number.asc())
    if start:
        q = q.where(Shipment.batch_date >= start)
    if end:
        q = q.where(Shipment.batch_date <= end)
    return list(session.scalars(q).all())


_PHONE_FIELDS = {"sender_phone", "receiver_phone"}


def _clean_phone_str(v: str) -> str:
    """Remove spurious '.0' suffix added when Excel stores phone numbers as
    numeric cells and openpyxl reads them as Python floats (e.g. 99111437.0
    → 99111437).  Only strips the suffix when the remainder is purely digits
    (with an optional leading +/-)."""
    if not v:
        return v
    s = str(v).strip()
    if s.endswith(".0") and s[:-2].lstrip("+-").isdigit():
        return s[:-2]
    return s


def to_dict(s: Shipment) -> dict:
    """Plain dict with date/datetime serialized as ISO strings so the result
    is JSON-safe (used both by FastAPI responses and the Jinja2 template)."""
    out = {}
    for c in s.__table__.columns:
        v = getattr(s, c.name)
        if isinstance(v, dt.datetime):
            v = v.isoformat()
        elif isinstance(v, dt.date):
            v = v.isoformat()
        elif c.name in _PHONE_FIELDS and v:
            v = _clean_phone_str(str(v))
        out[c.name] = v
    return out


def _resolve_price(payload_formula: str, payload_price: float,
                    weight: float, declared_value: float) -> tuple[str, float]:
    """If formula given, evaluate; else use the explicit price."""
    f = (payload_formula or "").strip()
    if f:
        return f, evaluate_price(f, weight=weight, declared_value=declared_value)
    return "", float(payload_price or 0)


def _compute_total(weight: float, price: float, extra: float) -> float:
    """Total = weight × price + extra. Always recomputed."""
    return (float(weight or 0) * float(price or 0)) + float(extra or 0)


# --------------------------------------------------------------------------
# app
# --------------------------------------------------------------------------
app = FastAPI(title="Mon Freight CDP")
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))

# Cache-busting version string — changes on every deploy so browsers
# always load the latest CSS/JS rather than a cached copy.
_STATIC_VERSION = str(int(dt.datetime.now().timestamp()))
templates.env.globals["_v"] = _STATIC_VERSION

# --- security & operations add-ons --------------------------------------
# auth: login + SMS verification + session protection for every page/API
# backup: automatic daily backups, Google Drive sync, restore
from auth import init_auth                      # noqa: E402
from backup import init_backup                  # noqa: E402
from activity_log import init_activity_log, log_activity, router as activity_router  # noqa: E402
from sms import init_sms, router as sms_router  # noqa: E402
from shipment_prep import init_prep              # noqa: E402
init_auth(app, engine, templates)
init_backup(app, engine)
init_activity_log(engine)
init_sms(engine)
init_prep(app, engine, templates, Shipment)
app.include_router(activity_router)
app.include_router(sms_router)


@app.get("/api/health")
def health():
    """Self-check: report DB connectivity, schema, row count."""
    try:
        with engine.connect() as conn:
            if DB_URL.startswith("sqlite"):
                cols = [r[1] for r in conn.execute(
                    text("PRAGMA table_info(shipments)")).fetchall()]
            else:
                cols = []
            n = conn.execute(text("SELECT COUNT(*) FROM shipments")).scalar() or 0
        return {
            "ok": True,
            "db_url_kind": "sqlite" if DB_URL.startswith("sqlite") else "postgres",
            "schema_columns": cols,
            "shipment_count": n,
            "expected_columns": [c.name for c in Shipment.__table__.columns],
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


PAGE_SIZE = 15


def _filter_shipments(rows: list, q: str) -> list:
    """Filter Shipment rows in Python by sender/receiver name+phone or
    batch date (ISO). Case-insensitive substring match."""
    q = (q or "").strip()
    if not q:
        return rows
    # Try exact-batch-date match first
    try:
        d = dt.date.fromisoformat(q)
        return [r for r in rows if r.batch_date == d]
    except ValueError:
        pass
    needle = q.lower()
    out = []
    for r in rows:
        hay = " ".join(str(getattr(r, c) or "").lower() for c in (
            "sender_name", "sender_phone",
            "receiver_name", "receiver_phone",
            "mf_number",
        ))
        if needle in hay:
            out.append(r)
    return out


@app.get("/", response_class=HTMLResponse)
def index(request: Request,
          date: Optional[str] = None,
          start: Optional[str] = None,
          end: Optional[str] = None,
          q: Optional[str] = None,
          page: int = 1):
    today = dt.date.today()
    default_date = dt.date.fromisoformat(date) if date else today
    start_d = dt.date.fromisoformat(start) if start else None
    end_d = dt.date.fromisoformat(end) if end else None
    page = max(1, int(page or 1))
    with Session(engine) as s:
        all_rows = all_shipments(s, start_d, end_d)
        filtered = _filter_shipments(all_rows, q or "")
        # Sort: newest batch first, then ascending box_number within each batch.
        # Using negative ordinal for date so we can keep box_number ascending
        # without a second reverse pass (avoids putting BOX 15 before BOX 1).
        filtered.sort(key=lambda r: (-r.batch_date.toordinal(), r.box_number or 0))
        total = len(filtered)
        start_idx = (page - 1) * PAGE_SIZE
        page_rows = filtered[start_idx:start_idx + PAGE_SIZE]
        all_dates = list(s.scalars(
            select(Shipment.batch_date).distinct().order_by(Shipment.batch_date.desc())
        ).all())
    pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    return templates.TemplateResponse("index.html", {
        "request": request,
        "default_date": default_date.isoformat(),
        "shipments": [to_dict(r) for r in page_rows],
        "all_dates": [d.isoformat() for d in all_dates],
        "start_d": start_d.isoformat() if start_d else "",
        "end_d": end_d.isoformat() if end_d else "",
        "today": today.isoformat(),
        "search_q": q or "",
        "page": page,
        "pages": pages,
        "total_count": total,
        "page_size": PAGE_SIZE,
    })


@app.get("/api/stats")
def shipment_stats(start: Optional[str] = None, end: Optional[str] = None,
                   q: Optional[str] = None):
    sd = dt.date.fromisoformat(start) if start else None
    ed = dt.date.fromisoformat(end) if end else None
    with Session(engine) as s:
        rows = all_shipments(s, sd, ed)
        if q:
            rows = _filter_shipments(rows, q)
    total   = len(rows)
    weight  = sum(float(r.weight        or 0) for r in rows)
    value   = sum(float(r.declared_value or 0) for r in rows)
    freight = sum(float(r.total_aud     or 0) for r in rows)
    paid    = sum(float(r.total_aud     or 0) for r in rows if r.paid)
    return {"total": total, "weight": weight, "declared_value": value,
            "freight": freight, "paid": paid, "outstanding": freight - paid}


@app.get("/api/shipments")
def list_shipments(start: Optional[str] = None, end: Optional[str] = None,
                   q: Optional[str] = None, page: int = 0):
    sd = dt.date.fromisoformat(start) if start else None
    ed = dt.date.fromisoformat(end) if end else None
    with Session(engine) as s:
        rows = all_shipments(s, sd, ed)
        if q:
            rows = _filter_shipments(rows, q)
        if page > 0:
            rows.sort(key=lambda r: (-r.batch_date.toordinal(), r.box_number or 0))
            start_idx = (page - 1) * PAGE_SIZE
            rows = rows[start_idx:start_idx + PAGE_SIZE]
        return [to_dict(r) for r in rows]


@app.post("/api/shipments")
def create_shipment(payload: ShipmentIn, request: Request):
    formula, price = _resolve_price(
        payload.price_formula, payload.price_aud,
        payload.weight, payload.declared_value)
    total = _compute_total(payload.weight, price, payload.extra_charges)
    with Session(engine) as s:
        # --- box number: explicit (validate uniqueness) or auto-assign ---
        if payload.box_number is not None:
            try:
                box = int(payload.box_number)
            except (TypeError, ValueError):
                raise HTTPException(400,
                    f"Box number must be a positive whole number, got "
                    f"'{payload.box_number}'.")
            if box <= 0:
                raise HTTPException(400,
                    "Box number must be 1 or greater.")
            existing = s.scalar(
                select(Shipment).where(
                    Shipment.batch_date == payload.batch_date,
                    Shipment.box_number == box,
                )
            )
            if existing:
                raise HTTPException(400,
                    f"Box number {box} is already taken for "
                    f"{payload.batch_date.isoformat()} "
                    f"(MF {existing.mf_number}). "
                    f"Pick a different number or leave it blank for auto-assign.")
        else:
            box = next_box_for(s, payload.batch_date)

        mf = mf_for(payload.batch_date, box)
        data = payload.model_dump(exclude={
            "batch_date", "box_number", "price_formula", "price_aud"})
        ship = Shipment(
            batch_date=payload.batch_date,
            box_number=box,
            mf_number=mf,
            price_formula=formula,
            price_aud=price,
            total_aud=total,
            **data,
        )
        s.add(ship)
        try:
            s.commit()
        except Exception as e:
            # Catch DB-level unique-constraint violations on mf_number.
            s.rollback()
            raise HTTPException(400,
                f"Could not save shipment — MF {mf} likely already exists "
                f"in the database. ({e.__class__.__name__})")
        s.refresh(ship)
        result = to_dict(ship)
        log_activity(_username(request), "shipment_created",
                     f"MF: {ship.mf_number}, batch: {ship.batch_date}, "
                     f"sender: {ship.sender_name}, receiver: {ship.receiver_name}",
                     _client_ip(request))
        return result


@app.post("/api/shipments/bulk-create")
def bulk_create_shipments(payload: BulkCreateIn, request: Request):
    """Create N blank shipments for a batch date so staff can record several
    boxes from the same customer first and fill in details afterwards.
    Box numbers are auto-assigned sequentially from the next free number."""
    count = payload.count
    if not isinstance(count, int) or count < 1:
        raise HTTPException(400, "Number of shipments must be a whole number of 1 or more.")
    if count > 50:
        raise HTTPException(400, "You can create at most 50 blank shipments at once.")
    created = []
    with Session(engine) as s:
        start_box = next_box_for(s, payload.batch_date)
        for i in range(count):
            box = start_box + i
            ship = Shipment(
                batch_date=payload.batch_date,
                box_number=box,
                mf_number=mf_for(payload.batch_date, box),
                price_aud=18.00,
            )
            s.add(ship)
            created.append(ship)
        try:
            s.commit()
        except Exception as e:
            s.rollback()
            raise HTTPException(400,
                f"Could not create blank shipments — a box or MF number may "
                f"already exist for {payload.batch_date.isoformat()}. "
                f"({e.__class__.__name__})")
        for ship in created:
            s.refresh(ship)
        result = [to_dict(sh) for sh in created]
        log_activity(_username(request), "shipments_bulk_created",
                     f"count: {len(created)}, batch: {payload.batch_date}, "
                     f"boxes: {created[0].box_number}–{created[-1].box_number}",
                     _client_ip(request))
        return result


@app.put("/api/shipments/{ship_id}")
def update_shipment(ship_id: int, payload: ShipmentIn, request: Request):
    with Session(engine) as s:
        ship = s.get(Shipment, ship_id)
        if not ship:
            raise HTTPException(404, "Shipment not found")

        date_changed = payload.batch_date != ship.batch_date
        new_box = payload.box_number  # may be None → keep existing/auto
        if date_changed:
            ship.batch_date = payload.batch_date
        if new_box is not None:
            try:
                box = int(new_box)
            except (TypeError, ValueError):
                raise HTTPException(400,
                    f"Box number must be a positive whole number, got '{new_box}'.")
            if box <= 0:
                raise HTTPException(400, "Box number must be 1 or greater.")
            if box != ship.box_number or date_changed:
                clash = s.scalar(
                    select(Shipment).where(
                        Shipment.batch_date == ship.batch_date,
                        Shipment.box_number == box,
                        Shipment.id != ship.id,
                    )
                )
                if clash:
                    raise HTTPException(400,
                        f"Box number {box} is already taken for "
                        f"{ship.batch_date.isoformat()} (MF {clash.mf_number}). "
                        f"Pick a different number.")
            ship.box_number = box
        elif date_changed:
            ship.box_number = next_box_for(s, ship.batch_date)
        # Always re-derive MF from the (possibly updated) box + date so the
        # last 3 digits match box_number.
        ship.mf_number = mf_for(ship.batch_date, ship.box_number)

        formula, price = _resolve_price(
            payload.price_formula, payload.price_aud,
            payload.weight, payload.declared_value)
        total = _compute_total(payload.weight, price, payload.extra_charges)
        for key, val in payload.model_dump(exclude={
                "batch_date", "box_number", "price_formula", "price_aud"}).items():
            setattr(ship, key, val)
        ship.price_formula = formula
        ship.price_aud = price
        ship.total_aud = total
        try:
            s.commit()
        except Exception as e:
            s.rollback()
            raise HTTPException(400,
                f"Could not save — MF {ship.mf_number} likely conflicts. "
                f"({e.__class__.__name__})")
        s.refresh(ship)
        result = to_dict(ship)
        log_activity(_username(request), "shipment_updated",
                     f"MF: {ship.mf_number}, batch: {ship.batch_date}, "
                     f"sender: {ship.sender_name}, receiver: {ship.receiver_name}",
                     _client_ip(request))
        return result


@app.patch("/api/shipments/{ship_id}")
def patch_shipment(ship_id: int, payload: ShipmentPatch, request: Request):
    """Partial update used by inline cell edits."""
    with Session(engine) as s:
        ship = s.get(Shipment, ship_id)
        if not ship:
            raise HTTPException(404, "Shipment not found")
        data = payload.model_dump(exclude_none=True)
        had_box_change = "box_number" in data

        date_changed = ("batch_date" in data
                        and data["batch_date"] != ship.batch_date)
        if date_changed:
            ship.batch_date = data.pop("batch_date")
        # Box number can be edited inline too. Validate & detect conflicts.
        if had_box_change:
            box = int(data.pop("box_number"))
            if box <= 0:
                raise HTTPException(400, "Box number must be 1 or greater.")
            if box != ship.box_number or date_changed:
                clash = s.scalar(
                    select(Shipment).where(
                        Shipment.batch_date == ship.batch_date,
                        Shipment.box_number == box,
                        Shipment.id != ship.id,
                    )
                )
                if clash:
                    raise HTTPException(400,
                        f"Box number {box} is already taken for "
                        f"{ship.batch_date.isoformat()} (MF {clash.mf_number}).")
            ship.box_number = box
        elif date_changed:
            ship.box_number = next_box_for(s, ship.batch_date)
        if date_changed or had_box_change:
            ship.mf_number = mf_for(ship.batch_date, ship.box_number)

        if "price_formula" in data:
            f = data.pop("price_formula").strip()
            if f:
                try:
                    ship.price_aud = evaluate_price(
                        f, weight=ship.weight, declared_value=ship.declared_value)
                except FormulaError as e:
                    raise HTTPException(400, f"Bad formula: {e}")
                ship.price_formula = f
            else:
                ship.price_formula = ""
        for key, val in data.items():
            setattr(ship, key, val)
        # If a formula exists and weight/value changed, re-evaluate price.
        if ship.price_formula and ("weight" in data or "declared_value" in data):
            try:
                ship.price_aud = evaluate_price(
                    ship.price_formula, weight=ship.weight,
                    declared_value=ship.declared_value)
            except FormulaError:
                pass
        # Total is always recomputed from weight × price + extra.
        ship.total_aud = _compute_total(
            ship.weight, ship.price_aud, ship.extra_charges)
        s.commit()
        s.refresh(ship)
        fields_changed = ", ".join(payload.model_dump(exclude_none=True).keys())
        log_activity(_username(request), "shipment_patched",
                     f"MF: {ship.mf_number}, fields: {fields_changed}",
                     _client_ip(request))
        return to_dict(ship)


@app.delete("/api/shipments/{ship_id}")
def delete_shipment(ship_id: int, request: Request):
    with Session(engine) as s:
        ship = s.get(Shipment, ship_id)
        if not ship:
            raise HTTPException(404, "Shipment not found")
        mf_num = ship.mf_number
        batch  = ship.batch_date
        s.delete(ship)
        s.commit()
        log_activity(_username(request), "shipment_deleted",
                     f"MF: {mf_num}, batch: {batch}", _client_ip(request))
        return {"ok": True}


@app.post("/api/shipments/bulk-delete")
def bulk_delete_shipments(payload: IdsPayload, request: Request):
    """Delete selected shipments in one action."""
    ids = sorted({int(i) for i in payload.ids if int(i) > 0})
    if not ids:
        raise HTTPException(400, "No shipment IDs provided")
    with Session(engine) as s:
        ships = list(s.scalars(select(Shipment).where(Shipment.id.in_(ids))).all())
        mf_list = ", ".join(sh.mf_number for sh in ships)
        for ship in ships:
            s.delete(ship)
        s.commit()
    log_activity(_username(request), "shipments_bulk_deleted",
                 f"Deleted {len(ships)} shipment(s): {mf_list[:500]}",
                 _client_ip(request))
    return {"ok": True, "deleted": len(ships)}


@app.post("/api/shipments/link")
def link_shipments(payload: IdsPayload):
    """Assign selected shipments to the same link group.

    If any selected shipment already belongs to a group, all members of
    that group are merged into one group (union).  The group id is the
    minimum shipment id across all members.
    """
    if len(payload.ids) < 2:
        raise HTTPException(400, "Select at least 2 shipments to link.")
    ids = list({int(i) for i in payload.ids})
    with Session(engine) as s:
        ships = list(s.scalars(select(Shipment).where(Shipment.id.in_(ids))).all())
        if len(ships) < 2:
            raise HTTPException(404, "One or more shipments not found.")
        # Collect existing groups to merge
        existing_groups = {sh.link_group for sh in ships if sh.link_group is not None}
        if existing_groups:
            group_members = list(s.scalars(
                select(Shipment).where(Shipment.link_group.in_(list(existing_groups)))
            ).all())
            # Union of selected + all existing group members
            all_ids = {sh.id for sh in ships} | {sh.id for sh in group_members}
            ships = list(s.scalars(select(Shipment).where(Shipment.id.in_(list(all_ids)))).all())
        new_group = min(sh.id for sh in ships)
        for sh in ships:
            sh.link_group = new_group
        s.commit()
        return {"ok": True, "link_group": new_group, "count": len(ships)}


@app.post("/api/shipments/unlink")
def unlink_shipments(payload: IdsPayload):
    """Remove selected shipments from their link group.

    If unlinking leaves only one member in a group, that lone member is
    also unlinked (a group of 1 has no meaning).
    """
    ids = list({int(i) for i in payload.ids})
    with Session(engine) as s:
        ships = list(s.scalars(select(Shipment).where(Shipment.id.in_(ids))).all())
        affected_groups = {sh.link_group for sh in ships if sh.link_group is not None}
        if not affected_groups:
            return {"ok": True, "count": 0}
        for sh in ships:
            sh.link_group = None
        # If a group now has only 1 member remaining, clear that one too
        for gid in affected_groups:
            remaining = list(s.scalars(
                select(Shipment).where(
                    Shipment.link_group == gid,
                    Shipment.id.not_in(ids)
                )
            ).all())
            if len(remaining) == 1:
                remaining[0].link_group = None
        s.commit()
        return {"ok": True, "count": len(ships)}


@app.get("/api/links")
def list_links():
    """Return every link group and its member boxes, independent of paging or
    the current filter, so the UI can show a complete and consistent
    'linked with' indicator next to each box number throughout the system.

    Shape: { "<group_id>": [{id, box_number, batch_date, receiver_name}, ...] }
    Only shipments that actually belong to a link group are included.
    """
    with Session(engine) as s:
        rows = list(s.scalars(
            select(Shipment)
            .where(Shipment.link_group.is_not(None))
            .order_by(Shipment.link_group, Shipment.box_number)
        ).all())
    groups: dict[int, list] = {}
    for r in rows:
        groups.setdefault(r.link_group, []).append({
            "id": r.id,
            "box_number": r.box_number,
            "batch_date": r.batch_date.isoformat(),
            "receiver_name": r.receiver_name or "",
        })
    return groups


@app.get("/api/dashboard")
def dashboard_summary():
    """Dashboard KPI figures — this month, all-time, and last 8 batches."""
    today = dt.date.today()
    first_this_month = today.replace(day=1)
    with Session(engine) as s:
        rows = all_shipments(s)

    def totals(rs):
        return {
            "shipments": len(rs),
            "weight": round(sum(float(r.weight or 0) for r in rs), 2),
            "declared_value": round(sum(float(r.declared_value or 0) for r in rs), 2),
            "revenue": round(sum(float(r.total_aud or 0) for r in rs), 2),
            "paid": round(sum(float(r.total_aud or 0) for r in rs if r.paid), 2),
            "unpaid": round(sum(float(r.total_aud or 0) for r in rs if not r.paid), 2),
            "paid_count": sum(1 for r in rs if r.paid),
            "unpaid_count": sum(1 for r in rs if not r.paid),
        }

    month_rows = [r for r in rows if r.batch_date >= first_this_month]
    latest_batches = []
    for d in sorted({r.batch_date for r in rows}, reverse=True)[:10]:
        br = [r for r in rows if r.batch_date == d]
        t = totals(br)
        t["date"] = d.isoformat()
        latest_batches.append(t)

    return {
        "all_time": totals(rows),
        "this_month": totals(month_rows),
        "latest_batches": latest_batches,
        "today": today.isoformat(),
        "this_month_label": today.strftime("%B %Y"),
    }


@app.get("/api/reports")
def reports_summary(
    start: Optional[str] = None,
    end: Optional[str] = None,
):
    """Detailed batch-by-batch report for the Reports panel.
    Returns per-batch breakdown within the requested date range."""
    sd = dt.date.fromisoformat(start) if start else None
    ed = dt.date.fromisoformat(end) if end else None
    with Session(engine) as s:
        rows = all_shipments(s, sd, ed)

    if not rows:
        return {"batches": [], "summary": {
            "shipments": 0, "weight": 0, "revenue": 0,
            "paid": 0, "unpaid": 0, "batches": 0,
        }}

    batches_map: dict[str, list] = {}
    for r in rows:
        key = r.batch_date.isoformat()
        batches_map.setdefault(key, []).append(r)

    batches = []
    for date_str in sorted(batches_map, reverse=True):
        rs = batches_map[date_str]
        revenue = sum(float(r.total_aud or 0) for r in rs)
        paid = sum(float(r.total_aud or 0) for r in rs if r.paid)
        batches.append({
            "date": date_str,
            "shipments": len(rs),
            "weight": round(sum(float(r.weight or 0) for r in rs), 2),
            "declared_value": round(sum(float(r.declared_value or 0) for r in rs), 2),
            "revenue": round(revenue, 2),
            "paid": round(paid, 2),
            "unpaid": round(revenue - paid, 2),
            "paid_count": sum(1 for r in rs if r.paid),
            "unpaid_count": sum(1 for r in rs if not r.paid),
        })

    total_rev = sum(b["revenue"] for b in batches)
    total_paid = sum(b["paid"] for b in batches)
    summary = {
        "batches": len(batches),
        "shipments": sum(b["shipments"] for b in batches),
        "weight": round(sum(b["weight"] for b in batches), 2),
        "declared_value": round(sum(b["declared_value"] for b in batches), 2),
        "revenue": round(total_rev, 2),
        "paid": round(total_paid, 2),
        "unpaid": round(total_rev - total_paid, 2),
    }
    return {"batches": batches, "summary": summary}


@app.get("/api/customers")
def customers_list(q: str = "", limit: int = 100):
    """Return deduplicated senders and receivers for the Customers panel."""
    q = (q or "").strip().lower()
    with Session(engine) as s:
        rows = all_shipments(s)
    out, seen = [], set()
    for r in rows:
        for side in ("sender", "receiver"):
            item = {
                "side": side,
                "name": getattr(r, f"{side}_name") or "",
                "phone": getattr(r, f"{side}_phone") or "",
                "address": getattr(r, f"{side}_address") or "",
                "city": getattr(r, f"{side}_city") or "",
                "country": getattr(r, f"{side}_country") or "",
                "last_batch": r.batch_date.isoformat(),
            }
            key = (side, item["name"].strip().lower(), item["phone"].strip())
            if key in seen or not (item["name"] or item["phone"]):
                continue
            hay = " ".join(str(v).lower() for v in item.values())
            if q and q not in hay:
                continue
            seen.add(key)
            out.append(item)
            if len(out) >= limit:
                return out
    return out


@app.get("/api/customers/search")
def customer_search(q: str = "", side: str = "sender", limit: int = 12):
    """Find distinct past customers (sender or receiver) by name, phone or
    city. Returns up to `limit` deduplicated customer-detail blobs, ordered
    so the most recently used customers come first."""
    q = (q or "").strip()
    # Require at least 3 characters before searching — keeps the dropdown
    # from popping up too eagerly while the user is still typing the
    # first few keystrokes of a new name.
    if len(q) < 3:
        return []
    if side not in ("sender", "receiver"):
        side = "sender"
    name_col = getattr(Shipment, f"{side}_name")
    phone_col = getattr(Shipment, f"{side}_phone")
    city_col = getattr(Shipment, f"{side}_city")
    field_names = [
        f"{side}_name", f"{side}_phone", f"{side}_address",
        f"{side}_city", f"{side}_country",
    ]
    if side == "sender":
        field_names.append("sender_postal")
    cols = [getattr(Shipment, f) for f in field_names]
    from sqlalchemy import or_
    pattern = f"%{q}%"
    with Session(engine) as s:
        rows = s.execute(
            select(*cols, Shipment.created_at)
            .where(or_(name_col.ilike(pattern),
                        phone_col.ilike(pattern),
                        city_col.ilike(pattern)))
            .order_by(Shipment.created_at.desc())
            .limit(300)
        ).all()
    seen, out = set(), []
    q_lower = q.lower()
    for r in rows:
        # First N elements are the field cols, last is created_at
        vals = r[:-1]
        key = (vals[0] or "").strip().lower(), (vals[1] or "").strip()
        if not any(key) or key in seen:
            continue
        seen.add(key)
        item = {fn: (v or "") for fn, v in zip(field_names, vals)}
        # Score: name-prefix match > phone-prefix > city-prefix > anywhere
        name_l = (item.get(f"{side}_name") or "").lower()
        phone_v = (item.get(f"{side}_phone") or "")
        city_l = (item.get(f"{side}_city") or "").lower()
        if name_l.startswith(q_lower):
            item["_score"] = 0
        elif phone_v.startswith(q):
            item["_score"] = 1
        elif city_l.startswith(q_lower):
            item["_score"] = 2
        else:
            item["_score"] = 3
        out.append(item)
        if len(out) >= limit * 3:  # over-fetch, then sort + trim
            break
    out.sort(key=lambda x: x["_score"])
    for x in out: x.pop("_score", None)
    return out[:limit]


@app.post("/api/preview-price")
def preview_price(payload: dict):
    """Live preview for the inline formula input."""
    try:
        v = evaluate_price(
            payload.get("formula", ""),
            weight=float(payload.get("weight") or 0),
            declared_value=float(payload.get("value") or 0),
        )
        return {"ok": True, "value": v}
    except FormulaError as e:
        return {"ok": False, "error": str(e)}


# --------------------------------------------------------------------------
# label printing
# --------------------------------------------------------------------------
def _label_context(d: dict) -> dict:
    bd = d.get("batch_date")
    if isinstance(bd, (dt.date, dt.datetime)):
        bd = bd.isoformat()
    return {
        "s": d,
        "description_clean": _strip_quantities(d.get("description") or ""),
        "receiver_full_address": _format_receiver_address(
            d.get("receiver_city"), d.get("receiver_address")),
        "batch_date_str": bd or "",
    }


@app.get("/shipments/{ship_id}/label", response_class=HTMLResponse)
def label_print(request: Request, ship_id: int):
    with Session(engine) as s:
        ship = s.get(Shipment, ship_id)
        if not ship:
            raise HTTPException(404, "Shipment not found")
        ctx = _label_context(to_dict(ship))
        return templates.TemplateResponse("label.html", {
            "request": request, "auto_print": True, **ctx,
        })


@app.get("/batches/{date}/labels.html", response_class=HTMLResponse)
def labels_print_all(request: Request, date: str):
    batch_date = dt.date.fromisoformat(date)
    with Session(engine) as s:
        rows = [to_dict(r) for r in shipments_for(s, batch_date)]
    return templates.TemplateResponse("labels_all.html", {
        "request": request,
        "labels": [_label_context(r) for r in rows],
        "auto_print": True,
    })


@app.get("/labels/by-ids", response_class=HTMLResponse)
def labels_by_ids(request: Request, ids: str = ""):
    """Render labels for an explicit list of shipment IDs.
    `ids` is a comma-separated list, e.g. /labels/by-ids?ids=3,7,12."""
    id_list = []
    for part in (ids or "").split(","):
        part = part.strip()
        if part.isdigit():
            id_list.append(int(part))
    if not id_list:
        raise HTTPException(400, "No shipment IDs provided")
    with Session(engine) as s:
        ships = list(s.scalars(
            select(Shipment).where(Shipment.id.in_(id_list))
        ).all())
    by_id = {sh.id: to_dict(sh) for sh in ships}
    ordered = [by_id[i] for i in id_list if i in by_id]
    if not ordered:
        raise HTTPException(404, "None of the requested shipments exist")
    return templates.TemplateResponse("labels_all.html", {
        "request": request,
        "labels": [_label_context(r) for r in ordered],
        "auto_print": True,
    })


# --------------------------------------------------------------------------
# Excel exports
# --------------------------------------------------------------------------
@app.get("/batches/{date}/aircargo.xlsx")
def export_aircargo(date: str, request: Request):
    try:
        batch_date = dt.date.fromisoformat(date)
    except ValueError:
        raise HTTPException(400, f"Bad date '{date}'. Use YYYY-MM-DD.")
    with Session(engine) as s:
        rows = [to_dict(r) for r in shipments_for(s, batch_date)]
    if not rows:
        raise HTTPException(404, f"No shipments recorded for {batch_date.isoformat()}")
    try:
        buf = io.BytesIO()
        build_aircargo_xlsx(rows, batch_date, buf)
        buf.seek(0)
    except Exception as e:
        raise HTTPException(500, f"Could not build Air Cargo .xlsx: {e}")
    log_activity(_username(request), "excel_export_aircargo",
                 f"Batch: {batch_date}, {len(rows)} shipment(s)",
                 _client_ip(request))
    fname = f"Air_Cargo_{batch_date.strftime('%d_%m_%Y')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{fname}"',
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
        },
    )


@app.get("/batches/{date}/labels.xlsx")
def export_labels(date: str, request: Request):
    try:
        batch_date = dt.date.fromisoformat(date)
    except ValueError:
        raise HTTPException(400, f"Bad date '{date}'. Use YYYY-MM-DD.")
    with Session(engine) as s:
        rows = [to_dict(r) for r in shipments_for(s, batch_date)]
    if not rows:
        raise HTTPException(404, f"No shipments recorded for {batch_date.isoformat()}")
    try:
        buf = io.BytesIO()
        build_labels_xlsx(rows, batch_date, buf)
        buf.seek(0)
    except FileNotFoundError as e:
        raise HTTPException(500, f"Label template missing: {e}")
    except Exception as e:
        raise HTTPException(500, f"Could not build labels .xlsx: {e}")
    log_activity(_username(request), "excel_export_labels",
                 f"Batch: {batch_date}, {len(rows)} shipment(s)",
                 _client_ip(request))
    fname = f"Labels_{batch_date.strftime('%d_%m_%Y')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{fname}"',
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
        },
    )


@app.get("/shipments/{ship_id}/label.xlsx")
def export_label_single(ship_id: int):
    """Download an Excel-formatted label for one shipment (uses the same
    label_template.xlsx layout as the batch export)."""
    with Session(engine) as s:
        ship = s.get(Shipment, ship_id)
        if not ship:
            raise HTTPException(404, "Shipment not found")
        row = to_dict(ship)
        batch_date = ship.batch_date
    buf = io.BytesIO()
    build_labels_xlsx([row], batch_date, buf)
    buf.seek(0)
    fname = f"Label_{row.get('mf_number') or ship_id}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{fname}"',
            # Bust any browser/proxy cache so deletes/edits show up
            # in the very next download.
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
        },
    )


@app.get("/labels/by-ids.xlsx")
def export_labels_by_ids(ids: str = ""):
    """Download Excel labels for an explicit comma-separated list of shipment
    IDs, e.g. /labels/by-ids.xlsx?ids=3,7,12. Uses the first shipment's
    batch date for the filename."""
    id_list = []
    for part in (ids or "").split(","):
        part = part.strip()
        if part.isdigit():
            id_list.append(int(part))
    if not id_list:
        raise HTTPException(400, "No shipment IDs provided")
    with Session(engine) as s:
        ships = list(s.scalars(
            select(Shipment).where(Shipment.id.in_(id_list))
        ).all())
    if not ships:
        raise HTTPException(404, "None of the requested shipments exist")
    by_id = {sh.id: sh for sh in ships}
    ordered = [by_id[i] for i in id_list if i in by_id]
    rows = [to_dict(sh) for sh in ordered]
    batch_date = ordered[0].batch_date
    buf = io.BytesIO()
    build_labels_xlsx(rows, batch_date, buf)
    buf.seek(0)
    fname = f"Labels_selected_{batch_date.strftime('%d_%m_%Y')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{fname}"',
            # Bust any browser/proxy cache so deletes/edits show up
            # in the very next download.
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
        },
    )


# --------------------------------------------------------------------------
# Excel-rendered HTML print views
#
# These endpoints take the SAME shipment data the .xlsx export would, build
# the workbook in memory, then render it as HTML using xlsx_html — which
# preserves borders, merges, fonts, column widths, and page breaks so the
# print pop-up looks identical to opening the .xlsx in Excel and printing.
# --------------------------------------------------------------------------
def _render_xlsx_labels_html(rows: list[dict], batch_date: dt.date,
                              title: str, scope: str) -> HTMLResponse:
    if not rows:
        raise HTTPException(404, "No shipments to render")
    from openpyxl import load_workbook
    buf = io.BytesIO()
    build_labels_xlsx(rows, batch_date, buf)
    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb.active

    # Limit the render to just the pages that hold real labels — each
    # printable page is `LABEL_PAGE_HEIGHT` rows tall and holds 2 labels
    # (left + right). For N labels we need ceil(N/2) pages, so the last
    # row to render is `pages * page_height + 1` (the +1 covers the
    # one-row spacer at the very top of the template).
    n = len(rows)
    pages_needed = (n + 1) // 2
    max_row = pages_needed * LABEL_PAGE_HEIGHT + 1

    html = render_print_page(ws, title=title, scope=scope,
                              page_height=LABEL_PAGE_HEIGHT,
                              max_row=max_row)
    return HTMLResponse(html, headers={
        "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
        "Pragma": "no-cache",
    })


@app.get("/shipments/{ship_id}/label.xlsx.html", response_class=HTMLResponse)
def export_label_single_html(ship_id: int):
    with Session(engine) as s:
        ship = s.get(Shipment, ship_id)
        if not ship:
            raise HTTPException(404, "Shipment not found")
        return _render_xlsx_labels_html(
            [to_dict(ship)], ship.batch_date,
            title=f"Label {ship.mf_number}",
            scope=f"BOX {ship.box_number} · {ship.mf_number}",
        )


@app.get("/labels/by-ids.xlsx.html", response_class=HTMLResponse)
def export_labels_by_ids_html(ids: str = ""):
    id_list = [int(p) for p in (ids or "").split(",")
               if p.strip().isdigit()]
    if not id_list:
        raise HTTPException(400, "No shipment IDs provided")
    with Session(engine) as s:
        ships = list(s.scalars(
            select(Shipment).where(Shipment.id.in_(id_list))
        ).all())
    if not ships:
        raise HTTPException(404, "None of the requested shipments exist")
    by_id = {sh.id: sh for sh in ships}
    ordered = [by_id[i] for i in id_list if i in by_id]
    rows = [to_dict(sh) for sh in ordered]
    return _render_xlsx_labels_html(
        rows, ordered[0].batch_date,
        title="Labels — selected",
        scope=f"{len(rows)} selected label{'' if len(rows) == 1 else 's'}",
    )


@app.get("/batches/{date}/labels.xlsx.html", response_class=HTMLResponse)
def export_labels_html(date: str):
    try:
        batch_date = dt.date.fromisoformat(date)
    except ValueError:
        raise HTTPException(400, f"Bad date '{date}'. Use YYYY-MM-DD.")
    with Session(engine) as s:
        rows = [to_dict(r) for r in shipments_for(s, batch_date)]
    return _render_xlsx_labels_html(
        rows, batch_date,
        title=f"Labels — {batch_date.isoformat()}",
        scope=f"{len(rows)} label(s) · batch {batch_date.isoformat()}",
    )


@app.post("/api/batches/{date}/import-aircargo")
async def import_aircargo(date: str, request: Request):
    """Bulk-import shipments from an existing Air Cargo .xlsx upload.

    The expected file layout matches what `build_aircargo_xlsx` produces:
      * Row 1-2: titles
      * Row 3:  metadata (date, MF batch id)
      * Row 4-5: section + column headers
      * Row 6+:  one BOX row per shipment, with column 1 starting "BOX ..."
        and columns 2-20 holding sender / receiver / cargo fields.

    Validation errors are surfaced as HTTP 400 with a human-readable
    `detail` string so the frontend can show a clear toast.
    """
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException

    # --- 1. validate batch date ---
    try:
        batch_date = dt.date.fromisoformat(date)
    except ValueError:
        raise HTTPException(400, f"Bad batch date '{date}'. Use YYYY-MM-DD.")

    # --- 2. read raw upload bytes ---
    body = await request.body()
    if not body:
        raise HTTPException(400, "Upload was empty. Pick a .xlsx file and try again.")
    if len(body) < 100:
        raise HTTPException(400,
            "File looks too small to be a valid .xlsx. "
            "Please re-export the Air Cargo workbook and try again.")

    # --- 3. parse workbook ---
    try:
        wb = load_workbook(io.BytesIO(body), data_only=True)
    except InvalidFileException:
        raise HTTPException(400,
            "That file is not a valid Excel workbook. "
            "Make sure you uploaded a .xlsx (not .xls, .csv, or .pdf).")
    except Exception as e:
        raise HTTPException(400, f"Could not read the .xlsx file: {e}")

    if not wb.sheetnames:
        raise HTTPException(400, "Workbook has no sheets — nothing to import.")
    ws = wb.active
    if ws.max_row < 1:
        raise HTTPException(400, f"Sheet '{ws.title}' appears to be empty.")

    def _num(v) -> float:
        """Best-effort numeric conversion; returns 0.0 on bad data."""
        if v is None or v == "":
            return 0.0
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0

    def _txt(v) -> str:
        return "" if v is None else str(v).strip()

    def _phone_txt(v) -> str:
        """Like _txt but also strips the '.0' suffix that openpyxl adds when
        a phone number is stored as a numeric cell in Excel."""
        return _clean_phone_str(_txt(v))

    import re
    _BOX_RE = re.compile(r"BOX\s*(\d+)", re.IGNORECASE)
    _MF_RE = re.compile(r"^MF\d{6,9}$", re.IGNORECASE)

    # Pre-load existing MF numbers so we can detect duplicates without
    # waiting for a database IntegrityError.
    with Session(engine) as s:
        existing_mfs = set(s.scalars(select(Shipment.mf_number)).all())

    added = 0
    skipped_rows = []   # rows with non-empty col-A that aren't BOX rows (logged for debug)
    errors = []
    seen_mfs_in_file = set()
    with Session(engine) as s:
        # Scan the ENTIRE sheet from row 1 so that files with fewer header
        # rows (e.g. 2 or 3 header rows instead of 5) don't lose early BOX rows.
        for r in range(1, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            if a is None or a == "":
                continue
            if not (isinstance(a, str) and a.strip().upper().startswith("BOX")):
                # Silently skip title / header / summary rows — do NOT count
                # as skipped so the user doesn't see a misleading "N rows skipped"
                # count for normal header content.
                continue
            try:
                # --- BOX number: parse "BOX N" from column A; fall back
                #     to next sequential if not parseable.
                m = _BOX_RE.search(a.strip())
                if m:
                    try:
                        box = int(m.group(1))
                    except ValueError:
                        box = next_box_for(s, batch_date)
                else:
                    box = next_box_for(s, batch_date)

                # --- MF number: read from column B. If blank, derive it
                #     from the chosen batch_date + box. Validate format
                #     and warn (without rejecting) if it looks malformed
                #     or doesn't match the chosen batch date.
                mf_raw = _txt(ws.cell(row=r, column=2).value)
                if mf_raw and mf_raw not in ("(auto)", "auto", "AUTO"):
                    mf = mf_raw.upper().replace(" ", "")
                    if not _MF_RE.match(mf):
                        errors.append(
                            f"row {r}: MF '{mf_raw}' doesn't look like "
                            f"MFYYMMDDNNN — saved as-is")
                    else:
                        # Surface a soft warning if the date encoded in
                        # the MF doesn't match the batch date the user
                        # picked in the import dialog.
                        date_in_mf = mf[2:8]
                        expected = batch_date.strftime("%y%m%d")
                        if date_in_mf != expected:
                            errors.append(
                                f"row {r}: MF '{mf}' encodes date "
                                f"{date_in_mf}, batch date is {expected} "
                                f"— saved as-is")
                else:
                    mf = mf_for(batch_date, box)

                # --- duplicate detection (existing DB row OR earlier
                #     row in this same file) ---
                if mf in existing_mfs:
                    errors.append(
                        f"row {r}: MF '{mf}' already exists in the "
                        f"database — skipping this row")
                    skipped_rows.append(r)
                    continue
                if mf in seen_mfs_in_file:
                    errors.append(
                        f"row {r}: MF '{mf}' appears more than once in "
                        f"this file — skipping the duplicate")
                    skipped_rows.append(r)
                    continue
                seen_mfs_in_file.add(mf)

                weight = _num(ws.cell(row=r, column=16).value)
                price = _num(ws.cell(row=r, column=17).value)
                extra = _num(ws.cell(row=r, column=18).value)
                ship = Shipment(
                    batch_date=batch_date,
                    box_number=box,
                    mf_number=mf,
                    sender_name=_txt(ws.cell(row=r, column=3).value),
                    sender_phone=_phone_txt(ws.cell(row=r, column=4).value),
                    sender_address=_txt(ws.cell(row=r, column=5).value),
                    sender_city=_txt(ws.cell(row=r, column=6).value),
                    sender_country=_txt(ws.cell(row=r, column=7).value) or "Австрали",
                    sender_postal=_txt(ws.cell(row=r, column=8).value),
                    receiver_name=_txt(ws.cell(row=r, column=9).value),
                    receiver_phone=_phone_txt(ws.cell(row=r, column=10).value),
                    receiver_address=_txt(ws.cell(row=r, column=11).value),
                    receiver_city=_txt(ws.cell(row=r, column=12).value),
                    receiver_country=_txt(ws.cell(row=r, column=13).value) or "Монгол",
                    description=_txt(ws.cell(row=r, column=14).value),
                    declared_value=_num(ws.cell(row=r, column=15).value),
                    weight=weight,
                    price_aud=price,
                    extra_charges=extra,
                    total_aud=(price * weight) + extra,
                    delivery_note=_txt(ws.cell(row=r, column=20).value),
                    notes=_txt(ws.cell(row=r, column=21).value),
                )
                s.add(ship)
                added += 1
            except Exception as row_err:
                errors.append(f"row {r}: {row_err}")
                continue
        s.commit()

    if added == 0:
        raise HTTPException(400,
            "No BOX rows were imported. "
            "Make sure column A contains 'BOX 1', 'BOX 2', etc. in the data rows "
            f"(scanned all {ws.max_row} rows). "
            + (f"Errors: {'; '.join(errors[:3])}" if errors else ""))

    log_activity(_username(request), "excel_import",
                 f"Batch: {date}, imported {added} shipment(s), "
                 f"skipped {len(skipped_rows)}"
                 + (f", errors: {'; '.join(errors[:3])}" if errors else ""),
                 _client_ip(request))
    return {
        "added": added,
        "skipped": len(skipped_rows),
        "errors": errors[:10],  # cap to avoid huge responses
    }
