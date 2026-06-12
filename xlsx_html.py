"""Render an openpyxl worksheet as HTML that visually matches Excel print
output as closely as possible.

Used by the "Excel label to print" option in the print-template chooser:
the browser pop-up shows the same layout the user would see if they
opened the .xlsx in Excel and printed it from there.

Translates:
  * column widths   (Excel chars   → mm)
  * row heights     (Excel points  → mm)
  * merged cells    (colspan / rowspan)
  * borders         (left/right/top/bottom + style)
  * fonts           (family, size, bold, color)
  * fills           (cell background)
  * alignment       (horizontal / vertical / wrap)
  * page breaks     (one printable page per `page_height` rows)
"""
from __future__ import annotations

from html import escape
from openpyxl.utils import get_column_letter

# Excel "character width" → millimetres. Excel measures column widths in
# units of "the average width of the digits 0-9 in the default font". For
# Calibri/Arial 11pt at 96 DPI this is ~7 px ≈ 1.85 mm.
EXCEL_CHAR_TO_MM = 1.83

# Excel row heights are in points; 1 point = 1/72 inch = 0.3528 mm.
POINT_TO_MM = 0.3528

# Default column width when none is set (Excel default is 8.43 chars).
DEFAULT_COL_WIDTH_CHARS = 8.43
# Default row height in points.
DEFAULT_ROW_HEIGHT_POINTS = 15.0


def _argb_to_hex(argb) -> str | None:
    """openpyxl colors are sometimes 'FFRRGGBB', sometimes objects.
    Return '#RRGGBB' or None if the colour can't be resolved."""
    if argb is None:
        return None
    s = str(argb)
    if len(s) == 8:
        return "#" + s[-6:]
    if len(s) == 6:
        return "#" + s
    return None


def _border_css(border) -> list[str]:
    """Convert openpyxl Border → list of CSS border-* declarations."""
    out = []
    for side in ("top", "right", "bottom", "left"):
        bs = getattr(border, side, None)
        if bs is None or not bs.style:
            continue
        # Map common Excel styles to CSS widths.
        width_map = {
            "thin": "1px", "hair": "1px", "dotted": "1px", "dashed": "1px",
            "medium": "2px", "thick": "3px",
            "double": "3px",
        }
        width = width_map.get(bs.style, "1px")
        style = "double" if bs.style == "double" else (
            "dashed" if bs.style in ("dashed", "dashDot") else (
            "dotted" if bs.style in ("dotted", "hair") else "solid"))
        color = "#000"
        if bs.color is not None:
            c = _argb_to_hex(bs.color.rgb)
            if c:
                color = c
        out.append(f"border-{side}:{width} {style} {color}")
    return out


def _font_css(font) -> list[str]:
    """openpyxl Font → CSS font-* declarations."""
    out = []
    if not font:
        return out
    if font.name:
        # Quote font names with spaces (e.g. "Code 128", "Arial Black").
        name = font.name
        if " " in name:
            name = f'"{name}"'
        out.append(f"font-family:{name},Arial,sans-serif")
    if font.size:
        # Excel font sizes are in points; CSS pt is the same unit.
        out.append(f"font-size:{float(font.size):.2f}pt")
    if font.bold:
        out.append("font-weight:700")
    if font.italic:
        out.append("font-style:italic")
    if font.color is not None:
        c = _argb_to_hex(font.color.rgb)
        if c and c.lower() != "#000000":
            out.append(f"color:{c}")
    return out


def _fill_css(fill) -> list[str]:
    """openpyxl PatternFill → background color (only 'solid' fills)."""
    if not fill or fill.patternType != "solid":
        return []
    fg = fill.fgColor
    c = _argb_to_hex(fg.rgb) if fg is not None else None
    if c and c.lower() not in ("#ffffff", "#000000"):
        return [f"background:{c}"]
    return []


def _align_css(alignment) -> list[str]:
    out = []
    if not alignment:
        return out
    if alignment.horizontal:
        h = alignment.horizontal
        if h in ("left", "right", "center", "justify"):
            out.append(f"text-align:{h}")
    if alignment.vertical:
        v_map = {"top": "top", "center": "middle", "bottom": "bottom"}
        out.append(f"vertical-align:{v_map.get(alignment.vertical, 'top')}")
    if alignment.wrap_text:
        out.append("white-space:normal;word-wrap:break-word;overflow-wrap:anywhere")
    else:
        out.append("white-space:nowrap;overflow:hidden")
    return out


def _build_merge_lookup(ws):
    """Return a dict (row, col) → tuple:
        ('start', rowspan, colspan) for top-left of a merge,
        ('skip',)                   for any other cell inside a merge.
    """
    lut = {}
    for rng in ws.merged_cells.ranges:
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                if r == rng.min_row and c == rng.min_col:
                    lut[(r, c)] = ("start",
                                    rng.max_row - rng.min_row + 1,
                                    rng.max_col - rng.min_col + 1)
                else:
                    lut[(r, c)] = ("skip",)
    return lut


def _format_cell_value(cell) -> str:
    """Stringify a cell value, respecting its Excel number format."""
    v = cell.value
    if v is None:
        return ""
    # Numbers that are clearly integers shouldn't show ".0".
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v)
    return escape(s).replace("\n", "<br>")


def render_worksheet_html(ws, page_height: int = 27,
                           page_break_every: bool = True,
                           max_row: int | None = None) -> str:
    """Render a worksheet as an HTML <table>. Per `page_height` rows insert
    a CSS page break so the print output paginates the same way the
    Excel file would.

    If `max_row` is set, only rows 1..max_row are rendered — useful when
    the worksheet is loaded from a template that has more empty slots than
    the caller actually wants to print (e.g. only 3 selected shipments
    in a 50-slot template)."""
    merge_lut = _build_merge_lookup(ws)
    last_row = ws.max_row if max_row is None else min(max_row, ws.max_row)

    # ----- column widths via <colgroup> -----
    col_widths_mm = []
    for c in range(1, ws.max_column + 1):
        letter = get_column_letter(c)
        dim = ws.column_dimensions.get(letter)
        w = dim.width if (dim and dim.width) else DEFAULT_COL_WIDTH_CHARS
        col_widths_mm.append(round(w * EXCEL_CHAR_TO_MM, 2))

    parts = ['<table class="xlsx-grid" cellspacing="0" cellpadding="0">']
    parts.append("<colgroup>")
    for w in col_widths_mm:
        parts.append(f'<col style="width:{w}mm">')
    parts.append("</colgroup>")
    parts.append("<tbody>")

    for r in range(1, last_row + 1):
        row_dim = ws.row_dimensions.get(r)
        h = row_dim.height if (row_dim and row_dim.height) else DEFAULT_ROW_HEIGHT_POINTS
        h_mm = round(h * POINT_TO_MM, 2)

        # Page break before every Nth row (so each label page sits on its
        # own A4 sheet). The first row is row 2 in the template (row 1 is
        # the small spacer at the very top).
        cls = ""
        if page_break_every and r > page_height + 1 and ((r - 2) % page_height == 0):
            cls = ' class="page-break-before"'

        parts.append(f'<tr style="height:{h_mm}mm"{cls}>')
        for c in range(1, ws.max_column + 1):
            mt = merge_lut.get((r, c))
            if mt and mt[0] == "skip":
                continue
            cell = ws.cell(row=r, column=c)

            attrs = []
            if mt and mt[0] == "start":
                _, rs, cs = mt
                if rs > 1: attrs.append(f'rowspan="{rs}"')
                if cs > 1: attrs.append(f'colspan="{cs}"')

            css = []
            if cell.has_style:
                css += _border_css(cell.border)
                css += _font_css(cell.font)
                css += _fill_css(cell.fill)
                css += _align_css(cell.alignment)
            else:
                css.append("vertical-align:top")
            # Tight default padding so the layout matches Excel's snug fit.
            css.append("padding:0 1px")

            style_attr = f' style="{";".join(css)}"' if css else ""
            attr_str = (" " + " ".join(attrs)) if attrs else ""
            parts.append(f"<td{attr_str}{style_attr}>{_format_cell_value(cell)}</td>")
        parts.append("</tr>")

    parts.append("</tbody></table>")
    return "".join(parts)


# Ready-made wrapper page so a route can return a complete HTML doc that
# prints itself when opened.
PRINT_PAGE_TEMPLATE = """<!DOCTYPE html>
<html lang="en"><head>
<meta charset="utf-8">
<title>{title}</title>
<style>
  @page {{ size: A4 landscape; margin: 0; }}
  * {{ box-sizing: border-box; }}
  html, body {{ margin: 0; padding: 0; background: #d8dde3; }}
  body {{ font-family: Arial, "Helvetica Neue", sans-serif; color: #000; }}

  .actions {{ text-align: center; padding: 10px; background: #f4f5f7; }}
  .actions button {{
    padding: 8px 18px; font-size: 14px;
    background: #1a505b; color: #fff; border: 0; border-radius: 5px;
    cursor: pointer; margin: 0 4px;
  }}
  .actions button.ghost {{
    background: #fff; color: #1a505b; border: 1px solid #1a505b;
  }}

  .sheet-wrap {{ padding: 8mm; display: flex; justify-content: center; }}
  .xlsx-grid {{
    border-collapse: collapse;
    table-layout: fixed;
    background: #fff;
    margin: 0 auto;
  }}
  .xlsx-grid td {{ overflow: hidden; }}
  .page-break-before {{ page-break-before: always; break-before: page; }}

  @media print {{
    body {{ background: #fff; }}
    .actions {{ display: none !important; }}
    .sheet-wrap {{ padding: 0; }}
    .xlsx-grid {{ width: 100%; }}
  }}
</style>
</head><body>
<div class="actions">
  <button onclick="window.print()">Print Excel labels</button>
  <button class="ghost" onclick="window.close()">Close</button>
  <span style="margin-left:14px;color:#555;font-size:12px;">{scope}</span>
</div>
<div class="sheet-wrap">{table_html}</div>
<script>
  window.addEventListener("load", () => setTimeout(() => window.print(), 600));
</script>
</body></html>"""


def render_print_page(ws, *, title: str, scope: str,
                       page_height: int = 27,
                       max_row: int | None = None) -> str:
    """Wrap render_worksheet_html() in a printable HTML page that
    auto-fires window.print() once it's loaded."""
    table_html = render_worksheet_html(ws, page_height=page_height,
                                         max_row=max_row)
    return PRINT_PAGE_TEMPLATE.format(
        title=escape(title), scope=escape(scope), table_html=table_html)
