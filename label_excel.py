"""
Excel exporters used by the web app:
  build_aircargo_xlsx — produces the same Air Cargo .xlsx layout
                        the team has been using.
  build_labels_xlsx   — produces the same multi-page label .xlsx,
                        ready to print from Excel.
"""
from __future__ import annotations

import datetime as dt
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent
LABEL_TEMPLATE = BASE_DIR / "data" / "label_template.xlsx"

# ---- shared with automate_labels.py: label layout constants ----
PAGE_HEIGHT = 27
LEFT_RIGHT_SHIFT = 25
SENDER_ADDRESS_PLACEHOLDER = (
    "Монголоос бусад Монголоос бусад Монголоос бусад "
)
AIMAG_PREFIXES = (
    "Архангай", "Баян-Өлгий", "Баянхонгор", "Булган", "Говь-Алтай",
    "Говьсүмбэр", "Дархан-Уул", "Дорноговь", "Дорнод", "Дундговь",
    "Завхан", "Орхон", "Өвөрхангай", "Өмнөговь", "Сэлэнгэ", "Төв",
    "Увс", "Ховд", "Хөвсгөл", "Хэнтий",
)
FIELDS = {
    "mf":                   (1, 4),
    "sender_name":          (3, 3),
    "receiver_name":        (3, 17),
    "sender_address":       (4, 3),
    "receiver_address":     (4, 17),
    "sender_country_code":  (6, 3),
    "sender_phone":         (6, 7),
    "receiver_country":     (6, 17),
    "receiver_phone":       (6, 21),
    "item_count":           (8, 1),
    "item_description":     (8, 2),
    "item_qty":             (8, 8),
    "item_value":           (8, 21),
    "item_currency":        (8, 24),
    "net_weight":           (10, 21),
    "postage":              (11, 24),
    "currency_label":       (13, 23),
    "mf_dup1":              (21, 21),
    "mf_dup2":              (22, 21),
    "date_mailed":          (24, 1),
    "box_label":            (25, 1),
}


def _xl(value):
    """Make a value safe to write into an .xlsx cell.

    Excel's SpreadsheetML forbids most ASCII control characters
    (\\x00-\\x08, \\x0b, \\x0c, \\x0e-\\x1f). openpyxl refuses to write them
    and raises IllegalCharacterError, which aborts the whole workbook
    build — one dirty character in one shipment kills the entire batch
    export. Operator-entered address/name/note fields occasionally pick
    these up from copy-paste or a bad encoding round-trip, so strip them
    at the point of writing rather than trusting the data.

    Non-string values (ints, floats, dates, None) pass through untouched.
    """
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value).strip()
    return value


def _strip_quantities(desc: str) -> str:
    if not desc:
        return ""
    return " ".join(t for t in str(desc).split() if not t.isdigit())


def _format_receiver_address(city: str, address: str) -> str:
    addr = (address or "").strip()
    city = (city or "").strip()
    if not addr:
        return city
    if not city or addr.startswith(city):
        return addr
    if "аймаг" in addr or any(addr.startswith(p) for p in AIMAG_PREFIXES):
        return addr
    return f"{city} {addr}"


def _label_top(k: int) -> int:
    return ((k - 1) // 2) * PAGE_HEIGHT + 2


def _col(k: int, left_col: int) -> int:
    return left_col + (0 if k % 2 == 1 else LEFT_RIGHT_SHIFT)


def _set(ws, k: int, field: str, value) -> None:
    row_off, left_col = FIELDS[field]
    row = _label_top(k) + row_off
    col = _col(k, left_col)
    # NOTE: ws.cell(row, col, value=None) is a no-op in openpyxl — it
    # only assigns when value is not None. To clear an existing cell we
    # must set the .value attribute directly. (This was the source of
    # stale-template-data showing up in every labels.xlsx export.)
    ws.cell(row=row, column=col).value = _xl(value)


# --------------------------------------------------------------------------
# Air Cargo export — builds a fresh workbook from scratch
# --------------------------------------------------------------------------
def build_aircargo_xlsx(rows: list[dict], batch_date: dt.date, out) -> None:
    """Write Air Cargo .xlsx matching the existing team format."""
    wb = Workbook()
    ws = wb.active
    ws.title = batch_date.strftime("%d%b%Y")

    bold = Font(bold=True, name="Arial")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="000000")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", start_color="DDEBF7")
    yellow = PatternFill("solid", start_color="FFF2CC")

    ws["A1"] = '"MON FREIGHT" PTY LTD to PPW (Taraaltad orson)'
    ws["A1"].font = Font(bold=True, size=14, name="Arial")
    ws.merge_cells("A1:T1")
    ws["A1"].alignment = center

    ws["A2"] = "АГААРЫН АЧААНЫ БҮРТГЭЛ"
    ws["A2"].font = Font(bold=True, size=12, name="Arial")
    ws.merge_cells("A2:T2")
    ws["A2"].alignment = center

    ws["N3"] = batch_date.strftime("%Y.%m.%d")
    ws["Q3"] = f"MF{batch_date.strftime('%Y%m%d')}"
    for c in ("N3", "Q3"):
        ws[c].font = bold

    sections = {
        "A4:B4": "№",
        "C4:H4": "Хэнээс",
        "I4:M4": "Хэнд",
        "N4:T4": "Ачааны тайлбар ",
    }
    for rng, val in sections.items():
        ws.merge_cells(rng)
        c = ws[rng.split(":")[0]]
        c.value = val
        c.font = bold
        c.alignment = center
        c.fill = header_fill

    headers = [
        "№", "MF дугаар ",
        "Овог, нэр", "Утас", "Гэрийн хаяг", "Хот", "Улс", "Шуудангийн дугаар",
        "Овог, нэр", "Утас", "Гэрийн хаяг", "Хот", "Улс",
        "Барааны тайлбар", "Нийт үнэлгээ", "Ачааны жин",
        "Нэгж үнэ AU$", "Нэмэлт төлбөр", "Нийт үнэ AU$",
        "Хүргэлттэй эсэх", "Дотоод тэмдэглэл",
    ]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = bold
        c.alignment = center
        c.fill = header_fill
        c.border = box

    paid_fill = PatternFill("solid", start_color="C6EFCE")  # Excel green

    # Link-group palette — 8 distinct soft colors cycling for each unique group
    _LINK_PALETTE = [
        "BDD7EE",  # cornflower blue
        "F4B8C1",  # rose pink
        "FFD966",  # amber
        "C9B1FF",  # lavender
        "92D050",  # lime (distinct from paid green C6EFCE)
        "F4A460",  # sandy orange
        "87CEEB",  # sky blue
        "DDA0DD",  # plum
    ]
    _sorted_groups = sorted({r.get("link_group") for r in rows
                              if r.get("link_group") is not None})
    _group_fill = {
        g: PatternFill("solid", start_color=_LINK_PALETTE[i % len(_LINK_PALETTE)])
        for i, g in enumerate(_sorted_groups)
    }

    rows = sorted(rows, key=lambda r: r["box_number"])
    for i, r in enumerate(rows):
        excel_row = 6 + i
        ws.cell(row=excel_row, column=1, value=_xl(f"BOX {r['box_number']}"))
        ws.cell(row=excel_row, column=2, value=_xl(r["mf_number"]))
        ws.cell(row=excel_row, column=3, value=_xl(r.get("sender_name") or ""))
        ws.cell(row=excel_row, column=4, value=_xl(r.get("sender_phone") or ""))
        ws.cell(row=excel_row, column=5, value=_xl(r.get("sender_address") or ""))
        ws.cell(row=excel_row, column=6, value=_xl(r.get("sender_city") or ""))
        ws.cell(row=excel_row, column=7, value=_xl(r.get("sender_country") or ""))
        ws.cell(row=excel_row, column=8, value=_xl(r.get("sender_postal") or ""))
        ws.cell(row=excel_row, column=9, value=_xl(r.get("receiver_name") or ""))
        ws.cell(row=excel_row, column=10, value=_xl(r.get("receiver_phone") or ""))
        ws.cell(row=excel_row, column=11, value=_xl(r.get("receiver_address") or ""))
        ws.cell(row=excel_row, column=12, value=_xl(r.get("receiver_city") or ""))
        ws.cell(row=excel_row, column=13, value=_xl(r.get("receiver_country") or ""))
        ws.cell(row=excel_row, column=14, value=_xl(r.get("description") or ""))
        ws.cell(row=excel_row, column=15, value=r.get("declared_value") or 0)
        ws.cell(row=excel_row, column=16, value=r.get("weight") or 0)
        ws.cell(row=excel_row, column=17, value=r.get("price_aud") or 0)
        ws.cell(row=excel_row, column=18, value=r.get("extra_charges") or 0)
        ws.cell(row=excel_row, column=19, value=r.get("total_aud") or 0)
        ws.cell(row=excel_row, column=20, value=_xl(r.get("delivery_note") or ""))
        ws.cell(row=excel_row, column=21, value=_xl(r.get("notes") or ""))
        is_paid = bool(r.get("paid"))
        lg = r.get("link_group")
        # Link group colour takes priority so all boxes in a group share
        # the same shade even when some are paid. Paid status is still
        # visible in the dedicated Status column (tick / green text).
        row_fill = _group_fill.get(lg) if lg is not None else (paid_fill if is_paid else None)
        for col in range(1, 22):
            cell = ws.cell(row=excel_row, column=col)
            cell.border = box
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if row_fill:
                cell.fill = row_fill

    widths = [7, 14, 22, 14, 30, 14, 14, 14, 22, 18, 30, 14, 14, 30, 12, 10, 12, 11, 12, 18, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[5].height = 36

    wb.save(out)


# --------------------------------------------------------------------------
# Label export — uses the bundled label_template.xlsx and overwrites cells
# --------------------------------------------------------------------------
def _populate_label(ws, k: int, r: dict, mailed_date: dt.datetime) -> None:
    mf_au = (r["mf_number"] + "AU") if r.get("mf_number") else ""
    _set(ws, k, "mf", mf_au)
    _set(ws, k, "sender_name", r.get("sender_name") or "")
    _set(ws, k, "receiver_name", r.get("receiver_name") or "")
    _set(ws, k, "sender_address", SENDER_ADDRESS_PLACEHOLDER)
    _set(ws, k, "receiver_address",
         _format_receiver_address(r.get("receiver_city"), r.get("receiver_address")))
    _set(ws, k, "sender_country_code", "AU")
    _set(ws, k, "sender_phone", "(+61)" + (str(r.get("sender_phone") or "").strip()))
    _set(ws, k, "receiver_country", "Монгол")
    _set(ws, k, "receiver_phone", str(r.get("receiver_phone") or "").strip())
    _set(ws, k, "item_count", 1)
    _set(ws, k, "item_description", _strip_quantities(r.get("description") or ""))
    _set(ws, k, "item_qty", 1)
    _set(ws, k, "item_value", r.get("declared_value") or 0)
    _set(ws, k, "item_currency", "AUD")
    _set(ws, k, "net_weight", r.get("weight") or 0)
    _set(ws, k, "postage", 10)
    _set(ws, k, "currency_label", "AUD")
    _set(ws, k, "mf_dup1", mf_au)
    _set(ws, k, "mf_dup2", mf_au)
    _set(ws, k, "date_mailed", mailed_date)
    _set(ws, k, "box_label", f"BOX {k}")


def _clear_label(ws, k: int) -> None:
    for f in FIELDS:
        _set(ws, k, f, None)


def _ensure_template_pages(ws, needed_slots: int) -> None:
    """Make sure the worksheet has enough label PAGES for `needed_slots`
    labels (max 150). Each page holds 2 labels (left + right). When the
    template is short, clone the first page's structure (cell formatting,
    merged cells, row heights, page break) onto fresh rows below.

    Performance note: merged cells are added via direct MergedCellRange
    construction (bypassing ws.merge_cells() which re-validates every
    contained cell) — this keeps 150-label generation under 10 seconds.
    """
    from openpyxl.worksheet.pagebreak import Break
    try:
        from openpyxl.worksheet.merge import MergedCellRange
    except ImportError:
        from openpyxl.worksheet.cell_range import CellRange as MergedCellRange

    PAGE_ROWS = PAGE_HEIGHT  # rows per label page in the template
    SRC_FIRST_ROW = 2        # first body row in the template
    SRC_LAST_ROW = SRC_FIRST_ROW + PAGE_ROWS - 1
    MAX_COL = ws.max_column or 50

    # How many pages does the template natively hold?
    current_pages = max((ws.max_row - SRC_FIRST_ROW + 1 + PAGE_ROWS - 1)
                        // PAGE_ROWS, 1)
    needed_pages = (needed_slots + 1) // 2

    if needed_pages <= current_pages:
        return

    # Snapshot the first page's structure so we can replicate it.
    # Store as (row_offset, min_col, row_offset2, max_col) tuples so we
    # can build range strings cheaply without repeated object creation.
    src_merges = []
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= SRC_FIRST_ROW and rng.max_row <= SRC_LAST_ROW:
            src_merges.append(
                (rng.min_row - SRC_FIRST_ROW, rng.min_col,
                 rng.max_row - SRC_FIRST_ROW, rng.max_col))

    src_row_heights = {}
    for r in range(SRC_FIRST_ROW, SRC_LAST_ROW + 1):
        rd = ws.row_dimensions.get(r)
        if rd is not None and rd.height is not None:
            src_row_heights[r - SRC_FIRST_ROW] = rd.height

    # Pre-snapshot styled cells to avoid re-reading ws on every page clone.
    # Each entry: (relative_row, col, _style, number_format | None)
    src_cells = []
    for r in range(SRC_FIRST_ROW, SRC_LAST_ROW + 1):
        for c in range(1, MAX_COL + 1):
            cell = ws.cell(row=r, column=c)
            if not cell.has_style:
                continue
            nf = cell.number_format if cell.number_format != "General" else None
            src_cells.append((r - SRC_FIRST_ROW, c, cell._style, nf))

    # Pre-build the existing row-break set so we don't add duplicates.
    existing_breaks = {b.id for b in ws.row_breaks.brk}

    # Clone every missing page.
    for page_idx in range(current_pages, needed_pages):
        offset = page_idx * PAGE_ROWS  # 0-based offset from SRC_FIRST_ROW
        abs_first = SRC_FIRST_ROW + offset

        # --- Copy cell styles ---
        for (rel_r, c, style, nf) in src_cells:
            dst = ws.cell(row=abs_first + rel_r, column=c)
            dst._style = style
            if nf:
                dst.number_format = nf

        # --- Re-apply merged ranges directly (fast path) ---
        # Constructing MergedCellRange objects directly avoids the O(N²)
        # ws.merge_cells() behaviour that visits every cell in every range.
        for (dr0, c0, dr1, c1) in src_merges:
            r0 = abs_first + dr0
            r1 = abs_first + dr1
            mcr = MergedCellRange(ws, f"{get_column_letter(c0)}{r0}:{get_column_letter(c1)}{r1}")
            ws.merged_cells.ranges.add(mcr)

        # --- Re-apply row heights ---
        for rel, h in src_row_heights.items():
            ws.row_dimensions[abs_first + rel].height = h

        # --- Insert page break (skip if already present) ---
        brk_id = abs_first - 1
        if brk_id not in existing_breaks:
            ws.row_breaks.append(Break(id=brk_id))
            existing_breaks.add(brk_id)


def build_labels_xlsx(rows: list[dict], batch_date: dt.date, out) -> None:
    if not LABEL_TEMPLATE.exists():
        raise FileNotFoundError(
            f"Label template not found at {LABEL_TEMPLATE}. "
            "Drop a Label_template.xlsx (a clean blank label workbook) "
            "into the data/ folder."
        )
    wb = load_workbook(str(LABEL_TEMPLATE))
    ws = wb.active

    rows = sorted(rows, key=lambda r: r["box_number"])

    # If the batch needs more label slots than the template ships with,
    # extend the template by cloning its first page. This is what makes
    # exports >50 print with proper borders, merges, and page-breaks
    # instead of just bare text.
    _ensure_template_pages(ws, len(rows))

    # Clear EVERY label slot in the template first, then populate only the
    # ones that correspond to real shipments. The template ships with
    # placeholder shipment data (50 sample rows); without this wipe, any
    # slot beyond `len(rows)` would leak old data into the download.
    n_slots_in_template = max(
        (ws.max_row + PAGE_HEIGHT - 2) // PAGE_HEIGHT, 1) * 2
    n_slots = max(n_slots_in_template, len(rows))
    for k in range(1, n_slots + 1):
        _clear_label(ws, k)

    mailed_dt = dt.datetime.combine(batch_date, dt.time())
    for i, r in enumerate(rows):
        # k must reflect the box's print order, starting at 1
        _populate_label(ws, i + 1, r, mailed_dt)

    wb.save(out)
